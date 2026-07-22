"""
excel_reporter.py â€“ Generates a rich Excel analysis report for Appium E2E results
===================================================================================
Creates an .xlsx workbook with:
  â€¢ Summary sheet  â€“ pass/fail counters, pie-chart-ready data, overall %
  â€¢ Details sheet  â€“ per-test row (name, module, status, duration, error, screenshot)
  â€¢ Analysis sheet â€“ grouped by module with colour-coded cells

Call reporter.add_result(â€¦) after each test, then reporter.save() at session end.
"""

import os
import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


# â”€â”€ Colour palette â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
CLR_PASS_BG   = "C6EFCE"
CLR_PASS_FG   = "276221"
CLR_FAIL_BG   = "FFC7CE"
CLR_FAIL_FG   = "9C0006"
CLR_SKIP_BG   = "FFEB9C"
CLR_SKIP_FG   = "9C6500"
CLR_HEADER_BG = "1A1A2E"
CLR_HEADER_FG = "E94560"
CLR_ALT_ROW   = "F5F5F5"
CLR_TITLE_BG  = "16213E"
CLR_TITLE_FG  = "0F3460"


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_font(bold=True):
    return Font(name="Calibri", bold=bold, color=CLR_HEADER_FG, size=11)


def _cell_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)


class ResultRecord:
    def __init__(
        self,
        test_id: str,
        name: str,
        module: str,
        status: str,            # "PASS" | "FAIL" | "SKIP"
        duration_sec: float,
        error_msg: str = "",
        screenshot_path: str = "",
        category: str = "Functional Testing",
    ):
        self.test_id        = test_id
        self.name           = name
        self.module         = module
        self.status         = status
        self.duration_sec   = duration_sec
        self.error_msg      = error_msg
        self.screenshot_path = screenshot_path
        self.category       = category
        self.timestamp      = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


class ExcelReporter:
    """Collects ResultRecord objects and writes a formatted Excel workbook."""

    def __init__(self):
        self._results: list[ResultRecord] = []
        self._start_time = datetime.datetime.now()

    # â”€â”€ Public API â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def add_result(self, result: ResultRecord):
        self._results.append(result)

    # â”€â”€ Save â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def save(self) -> str:
        ts       = self._start_time.strftime("%Y%m%d_%H%M%S")
        filename = f"devduel_e2e_report_{ts}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)

        wb = Workbook()
        self._build_summary_sheet(wb)
        self._build_details_sheet(wb)
        self._build_analysis_sheet(wb)

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(filepath)
        print(f"\nâœ…  Excel report saved â†’ {filepath}\n")
        return filepath

    # â”€â”€ Summary Sheet â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _build_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("ðŸ“Š Summary", 0)
        ws.sheet_view.showGridLines = False

        total = len(self._results)
        passed = sum(1 for r in self._results if r.status == "PASS")
        failed = sum(1 for r in self._results if r.status == "FAIL")
        skipped = total - passed - failed
        pct = round((passed / total) * 100, 1) if total else 0
        duration = (datetime.datetime.now() - self._start_time).total_seconds()

        # â”€â”€ Title banner â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        ws.merge_cells("A1:H1")
        cell = ws["A1"]
        cell.value = "ðŸŽ¯  DevDuel Mobile â€“ Appium E2E Test Report"
        cell.font      = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:H2")
        ws["A2"].value = f"Generated: {self._start_time.strftime('%d %B %Y  %H:%M')}   |   Total Duration: {duration:.1f}s"
        ws["A2"].font      = Font(name="Calibri", size=10, color="AAAAAA")
        ws["A2"].fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        # â”€â”€ KPI cards (row 4-8) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        kpis = [
            ("Total Tests",  total,   "1A1A2E", "FFFFFF"),
            ("âœ… Passed",     passed,  CLR_PASS_BG, CLR_PASS_FG),
            ("âŒ Failed",     failed,  CLR_FAIL_BG, CLR_FAIL_FG),
            ("âš ï¸ Skipped",    skipped, CLR_SKIP_BG, CLR_SKIP_FG),
            ("Pass Rate",    f"{pct}%", "0F3460", "FFFFFF"),
        ]
        col_start = 2
        for i, (label, value, bg, fg) in enumerate(kpis):
            col = col_start + i
            # Label row
            lc = ws.cell(row=4, column=col, value=label)
            lc.font      = Font(name="Calibri", bold=True, size=10, color=fg)
            lc.fill      = PatternFill("solid", fgColor=bg)
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.border    = _border()
            # Value row
            vc = ws.cell(row=5, column=col, value=value)
            vc.font      = Font(name="Calibri", bold=True, size=22, color=fg)
            vc.fill      = PatternFill("solid", fgColor=bg)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border    = _border()
            ws.row_dimensions[5].height = 50
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.row_dimensions[4].height = 22

        # â”€â”€ Pie chart data (hidden helper rows) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        data_row = 10
        ws.cell(row=data_row,     column=1, value="Status")
        ws.cell(row=data_row,     column=2, value="Count")
        ws.cell(row=data_row + 1, column=1, value="Passed")
        ws.cell(row=data_row + 1, column=2, value=passed)
        ws.cell(row=data_row + 2, column=1, value="Failed")
        ws.cell(row=data_row + 2, column=2, value=failed)
        ws.cell(row=data_row + 3, column=1, value="Skipped")
        ws.cell(row=data_row + 3, column=2, value=skipped)

        pie = PieChart()
        pie.title  = "Test Result Distribution"
        pie.style  = 10
        labels     = Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_row + 3)
        data_ref   = Reference(ws, min_col=2, min_row=data_row,     max_row=data_row + 3)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        ws.add_chart(pie, "D8")

        # â”€â”€ Column widths â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        ws.column_dimensions["A"].width = 16

    # â”€â”€ Details Sheet â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _build_details_sheet(self, wb: Workbook):
        ws = wb.create_sheet("📝 Test Details", 1)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"

        headers = [
            "#", "Test ID", "Test Name", "Module", "Category", "Status",
            "Duration (s)", "Timestamp", "Error Message", "Screenshot",
        ]
        col_widths = [5, 14, 45, 25, 25, 10, 14, 20, 50, 50]

        # Header rows
        ws.merge_cells("A1:J1")
        ws["A1"].value = "DevDuel – Detailed Test Results"
        ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor="E94560")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
        ws.row_dimensions[2].height = 20

        # Data rows
        for row_idx, result in enumerate(self._results, start=3):
            alt  = row_idx % 2 == 0
            fill = PatternFill("solid", fgColor=CLR_ALT_ROW if alt else "FFFFFF")

            if result.status == "PASS":
                status_fill = PatternFill("solid", fgColor=CLR_PASS_BG)
                status_font = Font(name="Calibri", bold=True, color=CLR_PASS_FG, size=10)
            elif result.status == "FAIL":
                status_fill = PatternFill("solid", fgColor=CLR_FAIL_BG)
                status_font = Font(name="Calibri", bold=True, color=CLR_FAIL_FG, size=10)
            else:
                status_fill = PatternFill("solid", fgColor=CLR_SKIP_BG)
                status_font = Font(name="Calibri", bold=True, color=CLR_SKIP_FG, size=10)

            row_data = [
                row_idx - 2,
                result.test_id,
                result.name,
                result.module,
                result.category,
                result.status,
                round(result.duration_sec, 2),
                result.timestamp,
                result.error_msg[:120] if result.error_msg else "",
                os.path.basename(result.screenshot_path),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border    = _border()
                cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
                if col_idx == 6:   # Status column
                    cell.fill = status_fill
                    cell.font = status_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.fill = fill
                    cell.font = Font(name="Calibri", size=10)
            ws.row_dimensions[row_idx].height = 18

        # Auto-filter
        ws.auto_filter.ref = f"A2:J{len(self._results) + 2}"

    # â”€â”€ Analysis Sheet â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _build_analysis_sheet(self, wb: Workbook):
        ws = wb.create_sheet("📊 Analysis", 2)
        ws.sheet_view.showGridLines = False

        # --- Section 1: Module-wise Test Analysis ---
        ws.merge_cells("A1:G1")
        ws["A1"].value = "Module-wise Test Analysis"
        ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Build module grouping
        modules: dict[str, dict] = {}
        for r in self._results:
            if r.module not in modules:
                modules[r.module] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0, "duration": 0.0}
            modules[r.module]["total"]    += 1
            modules[r.module]["duration"] += r.duration_sec
            if r.status == "PASS":
                modules[r.module]["passed"]  += 1
            elif r.status == "FAIL":
                modules[r.module]["failed"]  += 1
            else:
                modules[r.module]["skipped"] += 1

        # Header row
        headers = ["Module", "Total", "Passed", "Failed", "Skipped", "Pass Rate", "Avg Duration (s)"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor="E94560")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.row_dimensions[2].height = 20

        # Module rows
        for row_idx, (module, stats) in enumerate(modules.items(), start=3):
            pct = round((stats["passed"] / stats["total"]) * 100, 1) if stats["total"] else 0
            avg = round(stats["duration"] / stats["total"], 2)      if stats["total"] else 0

            row_data = [
                module,
                stats["total"],
                stats["passed"],
                stats["failed"],
                stats["skipped"],
                f"{pct}%",
                avg,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border    = _border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(name="Calibri", size=10)
                if col_idx == 3:
                    cell.fill = PatternFill("solid", fgColor=CLR_PASS_BG)
                elif col_idx == 4:
                    cell.fill = PatternFill("solid", fgColor=CLR_FAIL_BG)
                elif col_idx == 5:
                    cell.fill = PatternFill("solid", fgColor=CLR_SKIP_BG)
            ws.row_dimensions[row_idx].height = 18

        # Bar chart – pass / fail per module
        chart1 = BarChart()
        chart1.type    = "col"
        chart1.title   = "Pass vs Fail by Module"
        chart1.style   = 10
        chart1.y_axis.title = "Test Count"
        chart1.x_axis.title = "Module"
        chart1.grouping = "clustered"

        total_modules = len(modules)
        data1  = Reference(ws, min_col=3, max_col=4, min_row=2, max_row=2 + total_modules)
        cats1  = Reference(ws, min_col=1, min_row=3,             max_row=2 + total_modules)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.shape = 4
        ws.add_chart(chart1, "I2")

        # --- Section 2: Category-wise Test Analysis ---
        start_cat_row = 3 + total_modules + 2
        # Title row
        ws.merge_cells(f"A{start_cat_row}:G{start_cat_row}")
        title_cell = ws.cell(row=start_cat_row, column=1)
        title_cell.value = "Category-wise Test Analysis"
        title_cell.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        title_cell.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_cat_row].height = 30

        # Build category grouping
        categories: dict[str, dict] = {}
        for r in self._results:
            if r.category not in categories:
                categories[r.category] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0, "duration": 0.0}
            categories[r.category]["total"]    += 1
            categories[r.category]["duration"] += r.duration_sec
            if r.status == "PASS":
                categories[r.category]["passed"]  += 1
            elif r.status == "FAIL":
                categories[r.category]["failed"]  += 1
            else:
                categories[r.category]["skipped"] += 1

        # Header row
        header_cat_row = start_cat_row + 1
        headers_cat = ["Testing Category", "Total", "Passed", "Failed", "Skipped", "Pass Rate", "Avg Duration (s)"]
        for col_idx, h in enumerate(headers_cat, start=1):
            cell = ws.cell(row=header_cat_row, column=col_idx, value=h)
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor="0F3460")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()
        ws.row_dimensions[header_cat_row].height = 20

        # Category rows
        current_row = header_cat_row + 1
        for cat, stats in categories.items():
            pct = round((stats["passed"] / stats["total"]) * 100, 1) if stats["total"] else 0
            avg = round(stats["duration"] / stats["total"], 2)      if stats["total"] else 0

            row_data = [
                cat,
                stats["total"],
                stats["passed"],
                stats["failed"],
                stats["skipped"],
                f"{pct}%",
                avg,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border    = _border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(name="Calibri", size=10)
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_idx == 3:
                    cell.fill = PatternFill("solid", fgColor=CLR_PASS_BG)
                elif col_idx == 4:
                    cell.fill = PatternFill("solid", fgColor=CLR_FAIL_BG)
                elif col_idx == 5:
                    cell.fill = PatternFill("solid", fgColor=CLR_SKIP_BG)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # Category Chart
        chart2 = BarChart()
        chart2.type    = "col"
        chart2.title   = "Pass vs Fail by Testing Category"
        chart2.style   = 10
        chart2.y_axis.title = "Test Count"
        chart2.x_axis.title = "Category"
        chart2.grouping = "clustered"

        total_cats = len(categories)
        data2  = Reference(ws, min_col=3, max_col=4, min_row=header_cat_row, max_row=header_cat_row + total_cats)
        cats2  = Reference(ws, min_col=1, min_row=header_cat_row + 1,        max_row=header_cat_row + total_cats)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.shape = 4
        ws.add_chart(chart2, "I18")

